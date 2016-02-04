'''
@author: Jason Cheng
'''

from version import VERSION
from openpyxl import load_workbook
_version_ = VERSION

class OpenPyXlLibrary:


	ROBOT_LIBRARY_SCOPE = 'GLOBAL'
	ROBOT_LIBRARY_VERSION = VERSION

	def __init__(self):
		self.wb = None
		self.ws = None
		self.fileName = None

	def open_excel(self, filename):
		self.wb = load_workbook(filename)
		self.fileName = filename
		print self.fileName

	def put_string_to_cell(self, sheetname, column, row, value):
		self.ws = self.wb.get_sheet_by_name(sheetname)
		d = self.ws.cell(row = int(row), column = int(column))
		d.value = value

	def save_excel(self, filename):
		self.wb.save(filename)
	
	
	
	
	
	
	
	
	
	
	
	
	
	
	